import pandas as pd
from datetime import datetime
import numpy as np
import os
import glob, os
import pyodbc
import csv
import re
from openpyxl import load_workbook
import json
import re
personal_details = json.load(open('personal_details.json'))
server = personal_details["server"]
uid = personal_details["uid"]
pwd = personal_details["pwd"]

# cnt_forecast = pyodbc.connect('DRIVER={SQL Server Native Client 11.0};SERVER=;DATABASE=;UID=;')



def findFiles(folder_location):
    list_files_to_check = []
    dir_path = os.path.dirname(os.path.realpath(folder_location))
    for root, dirs, files in os.walk(dir_path):
        for file in files:
            if file.endswith('.xlsx'):
                list_files_to_check.append(root+'/'+str(file))

    for file_except in personal_details["file_exceptions"]:
        list_files_to_check = [x for x in list_files_to_check if file_except not in x]

    return (list_files_to_check)


# KRISHAN COPY ALL FILES TO ANOTHER LOCATION!!

def fileDupCheck(list_files_to_check):

    print ("blah blah")

   # for file in list_files_to_check:
   #      cursor = cnt_forecast.cursor()
   #      print (cursor)
   #      cursor.execute("INSERT INTO FileCheck (file_name) \
   #                 values (?)", "hello")
   #      cnt_forecast.commit()

def readData(list_files): # get data into 2 tables
    raw_data_list = []
    projectDets_table_list = []
    list_customers = []
    for file in list_files:

        date_created = (os.stat(file)[-1])
        date_modified = (os.stat(file)[-2])

        print (file)

        wb = load_workbook(file)
        sheet_names = (wb.sheetnames)
        df = pd.read_excel(open(file, 'rb'), sheet_name=sheet_names[0])
        projectDets_table = (df.loc[:,'Unnamed: 1':'Unnamed: 2']).head(6) # columns and rows filter.

        project_name_template = re.compile(r'PRO-\d\d\d\d')
        project_name = projectDets_table.loc[0, :'Unnamed: 2'][1][0:8]
        search_project_name = project_name_template.search(project_name)

        if search_project_name == None:
            pass
        if search_project_name != None:
            raw_data = ((df.loc[:,'Unnamed: 1':])).iloc[8:] # columns and rows filter.
            new_header = raw_data.iloc[0]  # grab the first row for the header
            raw_data = raw_data[2:]  # take the data less the header row
            raw_data.columns = new_header
            customer = os.path.split(os.path.dirname(file))[-1]
            list_customers.append(customer)
            raw_data_list.append(raw_data)
            projectDets_table_list.append(projectDets_table)

    return projectDets_table_list, raw_data_list,list_files,list_customers

def transformData(data):

    projectDets_table_list = data[0]
    raw_data_list = data[1]
    list_files = data[2]
    list_customers = data[3]
    first_upload_date_list = []
    refresh_date_list = []
    customer_list = []
    job_list = []
    agreed_project_cost_list = []
    discount_list = []
    status_list = []
    rate_group_list = []
    service_item_list = []
    employee_list = []
    item_rate_list = []
    week_start_date_list = []
    time_list = []
    output_files_list = []

    for projectDets_table,raw_data,files,customer in zip(projectDets_table_list,raw_data_list,list_files, list_customers):

        first_upload_date = datetime.now()
        refresh_date = datetime.now()
        job = projectDets_table.loc[0, :'Unnamed: 2'][1]
        est_project_cost = projectDets_table.loc[1, :'Unnamed: 2'][1]
        discount = projectDets_table.loc[2, :'Unnamed: 2'][1]
        agreed_project_cost = projectDets_table.loc[3,:'Unnamed: 2'][1]
        status = projectDets_table.loc[4, :'Unnamed: 2'][1]
        rate_group = projectDets_table.loc[5,:'Unnamed: 2'][1]

        dates = (list(raw_data.columns.values))
        len_dates = len(dates)
        dates = dates[6:len_dates]

        for date in dates:

            for index, row in raw_data.iterrows():

                service_item = row["Role"]
                employee =  row["Name"]
                item_rate = row["Rate"]
                week_start_date = date
                time = row[date]

                if employee is np.nan:
                    pass
                if employee is not np.nan:
                    if time is np.nan:
                        pass
                    if time is not np.nan:
                        transformed_data = (first_upload_date, refresh_date, customer, job, agreed_project_cost, discount, status, rate_group, service_item, employee, item_rate, week_start_date, time,files)

                        first_upload_date_list.append(first_upload_date)
                        refresh_date_list.append(refresh_date)
                        customer_list.append(customer)
                        job_list.append(job)
                        agreed_project_cost_list.append(agreed_project_cost)
                        discount_list.append(discount)
                        status_list.append(status)
                        rate_group_list.append(rate_group)
                        service_item_list.append(service_item)
                        employee_list.append(employee)
                        item_rate_list.append(item_rate)
                        week_start_date_list.append(week_start_date)
                        time_list.append(time)
                        output_files_list.append(files)

    frame = pd.DataFrame({"first_upload_date":first_upload_date_list,"refresh_date":refresh_date_list,"customer":customer_list,"job":job_list,"agreed_project_cost":agreed_project_cost_list,
            "discount":discount_list,"status":status_list,"rate_group":rate_group_list,"service_item":service_item_list,
            "employee":employee_list,"item_rate":item_rate_list,"week_start_date":week_start_date_list,"time":time_list,"file":output_files_list})

  #  print (frame)
    frame.to_csv("frame.csv")

                        # Writing to SQLSERVER
                        # cnt_achim = pyodbc.connect('DRIVER={SQL Server};SERVER=INSERT SERVER NAME HERE;DATABASE=INSERT DATABASE NAME HERE;UID=INSER USERNAME HERE;PWD=INSERT PASSWORD HERE')
                        # cursor = cnt_achim.cursor()
                        # cursor.execute("INSERT INTO ACHIM_SOCIAL_POSTS_FB (first_upload_date, refresh_date, customer, job, agreed_project_cost, discount, status, rate_group, service_item, employee, item_rate, week_start_date, time) \
                        #            values (?,?,?,?,?,?,?,?,?,?,?,?,?)", first_upload_date, refresh_date, customer, job, agreed_project_cost, discount, status, rate_group, service_item, employee, item_rate, week_start_date, time)
                        # cnt_achim.commit()

def updateChecker():
    print ("blah blah blah")


file_location = personal_details["file_location"]
list_files_to_check = findFiles(file_location)
data = readData(list_files_to_check)
transformed_data = transformData(data)
updateChecker()


# list_files = fileDupCheck(list_files_to_check)
# data = readData(list_files)
# transformed_data = transformData(data)
# updateChecker()